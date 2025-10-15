# app/reports.py
from __future__ import annotations  #  Debe ir en la primera línea

import datetime as dt
from zoneinfo import ZoneInfo
from openpyxl.chart.label import DataLabelList

from typing import List, Optional
from fastapi import Query, Body, APIRouter
from pydantic import BaseModel, EmailStr, Field, validator

# Stdlib
import os, json, unicodedata, re
from datetime import datetime, timedelta
from io import BytesIO

# Third-party
import requests
from flask import Blueprint, request, jsonify, make_response, Response

# DB y métricas (SIEMPRE desde services, no desde main)
from app.services.db import SessionLocal, ActionLog, ReportSchedule
from app.services.metrics import build_kpis, build_timeseries

# ERP opcional
try:
    from erp_client import list_unpaid_invoices
except Exception:
    list_unpaid_invoices = None

# Dependencias opcionales
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except Exception:
    Workbook = None
    Font = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
except Exception:
    canvas = None
    A4 = None

# ── Blueprint y constantes globales del módulo ───────────────────────────────
bp = Blueprint("reportes", __name__)  # este nombre se usa en main.register_blueprint
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "http://localhost:5050").rstrip("/")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Almacenes de configuración en archivos (simple, fácil de versionar)
GROUPS_JSON_PATH = os.getenv("REPORT_GROUPS_JSON_PATH", "data/report_groups.json")
os.makedirs(os.path.dirname(GROUPS_JSON_PATH) or ".", exist_ok=True)

RECIP_PATH = os.getenv("REPORT_RECIP_PATH", "report_recipients.json")

# Webhooks externos
N8N_TIMEOUT = int(os.getenv("N8N_TIMEOUT", "12") or 12)
N8N_REPORT_WEBHOOK_URL = os.getenv("N8N_REPORT_WEBHOOK_URL", "").strip()
EMAIL_WEBHOOK_URL = os.getenv("EMAIL_WEBHOOK_URL", "").strip()


# =============================================================================
# Helpers utilitarios
# =============================================================================

def _bool(v, default=False):
    """Parsea booleanos de query/body ('1','true','on' -> True)."""
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "yes", "y", "on")

def _split_csv(v):
    """Convierte 'a,b,c' -> ['a','b','c'] con limpieza básica."""
    if not v:
        return []
    if isinstance(v, list):
        return v
    return [s.strip() for s in str(v).split(",") if s and s.strip()]

def _csv_to_list(s: str) -> list[str]:
    return [x.strip() for x in (s or "").split(",") if x and x.strip()]

def _list_to_csv(v: list) -> str:
    return ",".join([str(x).strip() for x in (v or [])])

def _slugify(s: str) -> str:
    """'Dirección General' -> 'direccion-general' (para claves)."""
    s = unicodedata.normalize("NFD", s)
    s = s.encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", "-", s.strip().lower())
    return s

def _is_valid_email(s: str) -> bool:
    return bool(s) and EMAIL_RE.match(s.strip().lower())

def _dedup_valid(emails: list[str]) -> list[str]:
    """De-dup + validación simple de emails."""
    seen, out = set(), []
    for e in emails:
        e2 = (e or "").strip().lower()
        if EMAIL_RE.match(e2) and e2 not in seen:
            seen.add(e2)
            out.append(e2)
    return out

def _dedup(seq):
    """De-dup genérico para listas de strings."""
    seen = set()
    out = []
    for x in seq:
        x = (x or "").strip().lower()
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out

def _render_tpl(tpl: str, ctx: dict) -> str:
    """Reemplazo ultra-sencillo {{llaves}} en plantillas."""
    out = tpl or ""
    for k, v in ctx.items():
        out = out.replace(f"{{{{{k}}}}}", str(v))
    return out

def _report_links(period_days: int = 30) -> dict:
    """Enlaces públicos del backend para descargar el reporte."""
    base = PUBLIC_BASE_URL
    return {
        "pdf":  f"{base}/api/reports/summary.pdf?days={period_days}",
        "xlsx": f"{base}/api/reports/summary.xlsx?days={period_days}",
    }

from typing import List, Optional
from fastapi import Query, Body, APIRouter
from pydantic import BaseModel, EmailStr, Field, validator

router = APIRouter()

ALLOWED_FORMATS = {"pdf", "xlsx"}

def split_csv(value: Optional[str | List[str]]) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [p.strip() for p in str(value).split(",") if p.strip()]

class ReportParams(BaseModel):
    days: int = Field(30, ge=1, le=365)
    include_draft: bool = False
    formats: List[str] = Field(default_factory=lambda: ["pdf", "xlsx"])
    groups: List[str] = Field(default_factory=list)
    categories: List[str] = Field(default_factory=lambda: ["verde","amarillo","naranja","rojo"])
    emails: List[EmailStr] = Field(default_factory=list)
    min_dpd: int = Field(0, ge=0)

    @validator("formats", pre=True)
    def parse_formats(cls, v):
        vals = split_csv(v)
        vals = [s.lower() for s in vals]
        # valida formatos permitidos
        bad = [x for x in vals if x not in ALLOWED_FORMATS]
        if bad:
            raise ValueError(f"formatos no permitidos: {bad}")
        # de-dup
        return list(dict.fromkeys(vals))

    @validator("groups", pre=True)
    def parse_groups(cls, v): return split_csv(v)

    @validator("categories", pre=True)
    def parse_categories(cls, v): return split_csv(v)

    @validator("emails", pre=True)
    def parse_emails(cls, v): return split_csv(v)

@router.post("/api/report-schedule/run")
def run_report_post(payload: ReportParams = Body(...)):
    # payload llega tipado y validado
    return {"ok": True, "params": payload.dict()}

@router.get("/api/report-schedule/run")
def run_report_get(
    days: int = Query(30, ge=1, le=365),
    include_draft: bool = Query(False),
    formats: Optional[str] = Query(None, description="csv o repetir ?formats=pdf&formats=xlsx"),
    groups: Optional[str] = Query(None),
    categories: Optional[str] = Query("verde,amarillo,naranja,rojo"),
    emails: Optional[str] = Query(None),
    min_dpd: int = Query(0, ge=0),
):
    # Reusa el modelo para normalizar GET también
    params = ReportParams(
        days=days,
        include_draft=include_draft,
        formats=formats or ["pdf","xlsx"],
        groups=groups,
        categories=categories,
        emails=emails,
        min_dpd=min_dpd,
    )
    return {"ok": True, "params": params.dict()}


# =============================================================================
# Gestión de GRUPOS de correo (persistidos en JSON)
# =============================================================================

def _load_groups() -> dict:
    """Carga grupos → lista de correos desde archivo JSON."""
    if not os.path.exists(GROUPS_JSON_PATH):
        return {"direccion": [], "operaciones": [], "cobranza": [], "finanzas": []}
    try:
        with open(GROUPS_JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
            for k in ("direccion", "operaciones", "cobranza", "finanzas"):
                data.setdefault(k, [])
            return data
    except Exception:
        return {"direccion": [], "operaciones": [], "cobranza": [], "finanzas": []}

def _save_groups(groups: dict) -> None:
    """Guarda grupos → lista de correos en JSON."""
    with open(GROUPS_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(groups, f, ensure_ascii=False, indent=2)

def _normalize_groups_payload(payload: dict) -> dict:
    """
    Normaliza payloads como:
      { "groups": { "Dirección": ["a@x"], "cobranza": ["b@y"] } }
      o:
      { "Dirección": ["a@x"], "cobranza": ["b@y"] }
    -> { "direccion": [...], "cobranza": [...], ... }
    """
    g = payload.get("groups", payload)
    if not isinstance(g, dict):
        return {}
    norm = {}
    for k, v in g.items():
        key = _slugify(k)
        if not isinstance(v, list):
            continue
        cleaned = sorted({(e or "").strip() for e in v if isinstance(e, str) and "@" in e})
        norm[key] = cleaned
    for k in ("direccion", "operaciones", "cobranza", "finanzas"):
        norm.setdefault(k, [])
    return norm


@bp.get("/api/report-schedule/groups")
def api_get_report_groups():
    """Devuelve los grupos y sus correos (para UI administración)."""
    return jsonify({"ok": True, "groups": _load_groups()})

@bp.post("/api/report-schedule/groups")
def api_post_report_groups():
    """
    Mergea (suma) correos a los grupos existentes.
    Body: { "groups": { "Dirección": ["a@x"], "cobranza": ["b@y"] } }
    """
    body = request.get_json(silent=True) or {}
    incoming = _normalize_groups_payload(body)
    if not incoming:
        return jsonify({"ok": False, "error": "Body debe contener 'groups' como objeto {grupo:[emails]}"}), 400

    current = _load_groups()
    for key, emails in incoming.items():
        merged = sorted({*current.get(key, []), *emails})
        current[key] = merged

    _save_groups(current)
    return jsonify({"ok": True, "groups": current})


# =============================================================================
# Gestión de destinatarios (otra vista: config de recipientes)
# =============================================================================

def _load_recipients_config() -> dict:
    """
    Estructura de archivo:
    {
      "groups": {
        "direccion": [...],
        "operaciones": [...],
        "cobranza": [...],
        "finanzas": [...]
      }
    }
    Si no existe, se inicializa con listas vacías (o lo que haya en .env).
    """
    if os.path.exists(RECIP_PATH):
        with open(RECIP_PATH, "r", encoding="utf-8") as f:
            return json.load(f)

    cfg = {
        "groups": {
            "direccion":  _csv_to_list(os.getenv("REPORT_TO_DIRECCION", "")),
            "operaciones":_csv_to_list(os.getenv("REPORT_TO_OPERACIONES", "")),
            "cobranza":   _csv_to_list(os.getenv("REPORT_TO_COBRANZA", "")),
            "finanzas":   _csv_to_list(os.getenv("REPORT_TO_FINANZAS", "")),
        }
    }
    with open(RECIP_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return cfg

def _save_recipients_config(cfg: dict):
    with open(RECIP_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def compute_final_recipients(selected_groups: list[str], extras: list[str]) -> list[str]:
    """Une correos de grupos + extra_emails y normaliza."""
    cfg = _load_recipients_config()
    groups = cfg.get("groups", {})
    pool = []
    for g in selected_groups or []:
        pool += groups.get(g, [])
    pool += extras or []
    return _dedup_valid(pool)


@bp.get("/api/report-recipients")
def api_get_report_recipients():
    """Devuelve el JSON de recipientes (otra vista)."""
    return jsonify(_load_recipients_config()), 200

@bp.post("/api/report-recipients")
def api_save_report_recipients():
    """Guarda el JSON de recipientes (valida emails)."""
    data = request.get_json(force=True) or {}
    groups = (data.get("groups") or {})
    clean = {k: _dedup_valid(v or []) for k, v in groups.items()}
    cfg = {"groups": clean}
    _save_recipients_config(cfg)
    return jsonify(ok=True, config=cfg), 200

@bp.post("/api/report-recipients/preview")
def api_preview_recipients():
    """Vista previa: devuelve la lista final de correos a enviar."""
    data = request.get_json(force=True) or {}
    selected = data.get("groups") or []
    extras   = data.get("extras")  or []
    result = compute_final_recipients(selected, extras)
    return jsonify(ok=True, recipients=result, count=len(result)), 200


# =============================================================================
# Programación de reportes (modelo ReportSchedule en DB)
# =============================================================================

def _compute_next_run(rs: ReportSchedule) -> datetime:
    """Calcula el próximo envío (timezone-aware) y retorna dt con zona local."""
    try:
        tz = ZoneInfo(rs.timezone or "America/Mexico_City")
    except Exception:
        tz = ZoneInfo("America/Mexico_City")

    now = datetime.now(tz)
    hh, mm = (rs.time_of_day or "09:00").split(":")
    target = now.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)

    if (rs.frequency or "daily") == "daily":
        nxt = target if target > now else (target + timedelta(days=1))

    elif rs.frequency == "weekly":
        # lunes=0, domingo=6
        days = [int(x) for x in _csv_to_list(rs.days_of_week or "") if str(x).isdigit()]
        if not days:
            days = [0]
        cur = now.weekday()
        delta = min(((d - cur) % 7) or 7 for d in days)
        candidate = target + timedelta(days=delta)
        if candidate <= now:
            candidate = candidate + timedelta(days=7)
        nxt = candidate

    elif rs.frequency == "monthly":
        dom = max(1, min(int(rs.day_of_month or 1), 28))
        year, month = now.year, now.month
        candidate = now.replace(day=dom, hour=int(hh), minute=int(mm), second=0, microsecond=0)
        if candidate <= now:
            month = 1 if month == 12 else month + 1
            year = year + 1 if month == 1 else year
            candidate = candidate.replace(year=year, month=month)
        nxt = candidate

    else:
        nxt = target if target > now else (target + timedelta(days=1))

    return nxt

def _load_single_schedule(db) -> ReportSchedule:
    """Obtiene (o crea) un único registro de programación de reportes."""
    rs = db.query(ReportSchedule).order_by(ReportSchedule.id.asc()).first()
    if not rs:
        rs = ReportSchedule()
        db.add(rs)
        db.commit()
        db.refresh(rs)
    return rs

def send_report_to_n8n(payload: dict) -> dict:
    """
    Envía el reporte al webhook específico de REPORTES (no al webhook general).
    Estructura mínima esperada:
      {
        "emails": [...],
        "subject": "Reporte CxC",
        "body"/"html": "...",
        "links": { "pdf": "...", "xlsx": "..." }
      }
    """
    if not N8N_REPORT_WEBHOOK_URL:
        raise RuntimeError("N8N_REPORT_WEBHOOK_URL no configurado")
    r = requests.post(N8N_REPORT_WEBHOOK_URL, json=payload, timeout=N8N_TIMEOUT)
    try:
        return r.json()
    except Exception:
        return {"status_code": r.status_code, "text": r.text[:500]}

def _send_report_now(rs: ReportSchedule) -> dict:
    """
    Construye payload a partir del ReportSchedule, lo envía a n8n
    y actualiza last_run/next_run.
    """
    # Enlaces de descarga
    links = _report_links(rs.period_days or 30)
    fmts = set(_csv_to_list(rs.formats or "pdf"))

    # Render de asunto/cuerpo desde plantillas simples
    ctx = {
        "rango": rs.period_days or 30,
        "fecha": datetime.utcnow().strftime("%Y-%m-%d %H:%M"),
        "total": "-",  # si quieres, calcula totales reales y ponlos aquí
    }
    subject = _render_tpl(rs.subject_tpl or "Reporte ejecutivo CxC ({{rango}} días)", ctx)
    body    = _render_tpl(rs.body_tpl    or "Adjunto enlaces: {{fecha}}.", ctx)

    payload = {
        "type": "cxc_report",
        "period_days": rs.period_days or 30,
        "include_draft": bool(rs.include_draft),
        "categories": _csv_to_list(rs.categories or ""),
        "min_dpd": rs.min_dpd,
        "formats": list(fmts),
        "links": {k: v for k, v in links.items() if k in fmts},  # filtra formatos
        "groups": _csv_to_list(rs.groups or ""),
        "emails": _csv_to_list(rs.emails or ""),
        "subject": subject,
        "body": body,
    }

    # Envío
    resp = {}
    try:
        resp = send_report_to_n8n(payload)
    except Exception as e:
        resp = {"error": str(e)}

    # Auditoría y reprogramación
    with SessionLocal() as db:
        row = db.query(ReportSchedule).get(rs.id)
        if row:
            row.last_run = datetime.utcnow()
            nxt = _compute_next_run(row)
            row.next_run = nxt.astimezone(ZoneInfo("UTC")).replace(tzinfo=None)
            db.commit()

    return resp

# ====== Paleta y helpers para Board Directivo ======
from reportlab.lib import colors

PALETTE = {
    "ink": colors.HexColor("#111827"),
    "muted": colors.HexColor("#6B7280"),
    "line": colors.HexColor("#E5E7EB"),
    "card": colors.HexColor("#F8FAFC"),
    "blue": colors.HexColor("#2563EB"),
    "green": colors.HexColor("#10B981"),
    "yellow": colors.HexColor("#F59E0B"),
    "orange": colors.HexColor("#F97316"),
    "red": colors.HexColor("#EF4444"),
    "slate": colors.HexColor("#0f172a"),
}

def _safe_float(x) -> float:
    try:
        return float(x or 0.0)
    except Exception:
        return 0.0

def _days_past_due(due_date: str) -> int:
    """Convierte due_date (YYYY-MM-DD) a días de atraso (dpd ≥ -∞)."""
    try:
        d = dt.date.fromisoformat((due_date or "").strip())
        return (dt.date.today() - d).days
    except Exception:
        return 0

def _load_invoices_for_board(limit=1000):
    """
    Devuelve facturas desde ERP (solo impagas).
    Estructura mínima: [{customer, due_date, outstanding_amount}, ...]
    """
    if list_unpaid_invoices is None:
        return []

    rows = list_unpaid_invoices(
        limit=limit, include_draft=False, order_by="due_date asc"
    )
    return rows or []

def _board_metrics(rows: list[dict]) -> dict:
    """
    Calcula KPIs y buckets del tablero directivo.
    - on_time: dpd < 0
    - due_today: dpd == 0
    - late_1_14 / late_15_30 / late_31_60 / late_60_plus
    - en_tiempo (verde+amarillo) = on_time + due_today
    - riesgo_alto (>=15d)
    También calcula Top 5 clientes por saldo.
    """
    total = 0.0
    on_time = 0.0
    due_today = 0.0
    late_1_14 = late_15_30 = late_31_60 = late_60_plus = 0.0

    by_customer = {}

    for r in rows:
        amt = _safe_float(r.get("outstanding_amount"))
        total += amt
        dpd = _days_past_due(r.get("due_date"))
        cust = (r.get("customer") or "Cliente").strip()
        by_customer[cust] = by_customer.get(cust, 0.0) + amt

        if dpd < 0:
            on_time += amt
        elif dpd == 0:
            due_today += amt
        elif 1 <= dpd <= 14:
            late_1_14 += amt
        elif 15 <= dpd <= 30:
            late_15_30 += amt
        elif 31 <= dpd <= 60:
            late_31_60 += amt
        else:
            late_60_plus += amt

    en_tiempo = on_time + due_today
    riesgo_alto = late_15_30 + late_31_60 + late_60_plus

    top5 = sorted(by_customer.items(), key=lambda kv: kv[1], reverse=True)[:5]

    return {
        "total": total,
        "on_time": on_time,
        "due_today": due_today,
        "late_1_14": late_1_14,
        "late_15_30": late_15_30,
        "late_31_60": late_31_60,
        "late_60_plus": late_60_plus,
        "en_tiempo": en_tiempo,
        "riesgo_alto": riesgo_alto,
        "top5": [{"cliente": k, "monto": v} for k, v in top5],
    }


# =============================================================================
# Endpoints de previsualización y envío
# =============================================================================

@bp.post("/api/report-schedule/resolve")
def api_resolve_recipients():
    """
    Resuelve destinatarios a partir de grupos + extra_emails.
    Body:
      { "groups": ["direccion","operaciones"], "extra_emails": ["foo@bar.com"] }
    """
    body = request.get_json(silent=True) or {}
    selected = body.get("groups") or []
    extra    = body.get("extra_emails") or []

    groups = _load_groups()
    recipients = set(e for e in extra if isinstance(e, str) and "@" in e)
    for g in selected:
        key = _slugify(g)
        recipients.update(groups.get(key, []))

    recipients = sorted({e.strip().lower() for e in recipients if _is_valid_email(e)})
    return jsonify({"ok": True, "recipients": recipients})

@bp.post("/api/report-schedule/preview")
def api_report_preview():
    """
    Vista previa para la modal del front.
    Body:
      { "groups": ["direccion","cobranza"], "extra_emails": ["a@acme.com"], "period_days": 30 }
    """
    body = request.get_json(silent=True) or {}
    period_days   = int(body.get("period_days", 30))

    # Resolver destinatarios (sin llamar a otro endpoint)
    selected = body.get("groups") or []
    extras   = body.get("extra_emails") or []
    groups = _load_groups()
    recipients = set(e for e in extras if isinstance(e, str) and "@" in e)
    for g in selected:
        key = _slugify(g)
        recipients.update(groups.get(key, []))
    recipients = sorted({e.strip().lower() for e in recipients if _is_valid_email(e)})

    links = _report_links(period_days)
    return jsonify({
        "ok": True,
        "recipients": recipients,
        "links": links,
        "meta": {"period_days": period_days}
    })

@bp.post("/api/report-schedule/send-now")
def api_report_send_now():
    """
    Envía el reporte **ya** usando la programación guardada en DB (ReportSchedule).
    Query opcional: ?if_due=1 → solo envía si ya toca (según next_run).
    """
    only_if_due = (request.args.get("if_due") in ("1","true","yes","on"))
    with SessionLocal() as db:
        rs = _load_single_schedule(db)
        if only_if_due:
            if not rs.enabled:
                return jsonify({"ok": True, "sent": False, "reason": "disabled"})
            now_utc = datetime.utcnow()
            if rs.next_run and now_utc < rs.next_run:
                return jsonify({"ok": True, "sent": False, "reason": "not_due"})

        resp = _send_report_now(rs)
        return jsonify({"ok": True, "sent": True, "n8n": resp})


# =============================================================================
# Descargas del reporte: XLSX y PDF (a partir de KPIs de ActionLog)
# =============================================================================

@bp.get("/api/reports/summary.xlsx")
def report_summary_xlsx():
    """
    Resumen ejecutivo de acciones (KPIs + evolución) en XLSX.
    Incluye formato, barra de datos segura y gráfica de líneas.
    """
    # --- deps locales (evita romper si openpyxl no está) ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import LineChart, Reference
        from openpyxl.formatting.rule import DataBarRule
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify(ok=False, error="openpyxl no instalado"), 500

    # --- rango de fechas ---
    try:
        days = int(request.args.get("days", "30"))
    except ValueError:
        days = 30

    end_dt = datetime.utcnow()
    start_dt = end_dt - timedelta(days=days)

    # --- datos ---
    with SessionLocal() as db:
        kpis = build_kpis(db, start_dt, end_dt)
        ts   = build_timeseries(db, start_dt, end_dt)

    # Normaliza tasa de éxito (0..1); muchos backends la devuelven 0..100
    success = float(kpis.get("success_rate", 0.0) or 0.0)
    if success > 1.0:
        success = success / 100.0

    # --- libro y estilos ---
    wb = Workbook()
    ws = wb.active
    ws.title = "KPIs"

    bold   = Font(bold=True)
    h1     = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")
    thin   = Side(style="thin", color="DDDDDD")
    box    = Border(top=thin, left=thin, right=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="F3F4F6")  # gris claro

    # --- cabecera ---
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Resumen KPIs ({days} días)"
    ws["A1"].font = h1
    ws["A1"].alignment = center

    ws.append(["Inicio", start_dt.isoformat()])
    ws.append(["Fin",    end_dt.isoformat()])
    ws.append([])

    # --- tabla KPIs ---
    ws.append(["KPI", "Valor", "Descripción"])
    ws["A5"].font = ws["B5"].font = ws["C5"].font = bold
    ws["A5"].fill = ws["B5"].fill = ws["C5"].fill = head_fill
    ws["A5"].border = ws["B5"].border = ws["C5"].border = box

    rows_kpi = [
        ("Total acciones", kpis.get("total_actions", 0), ""),
        ("OK",             kpis.get("ok", 0),            ""),
        ("Fallidas",       kpis.get("failed", 0),        ""),
        ("Tasa de éxito",  success,                      "Porcentaje"),
    ]
    start_r = ws.max_row + 1
    for k, v, d in rows_kpi:
        ws.append([k, v, d])
        r = ws.max_row
        # formato por tipo
        if k == "Tasa de éxito":
            ws[f"B{r}"].number_format = "0.00%"
        else:
            ws[f"B{r}"].number_format = "0"
        # bordes
        ws[f"A{r}"].border = ws[f"B{r}"].border = ws[f"C{r}"].border = box
        ws[f"A{r}"].font = bold

    # Barra de datos (segura): sin end_value con end_type='max'
    try:
        ws.conditional_formatting.add(
            f"B{start_r}:B{start_r+2}",
            DataBarRule(
                start_type="num", start_value=0,
                end_type="max", end_value=None,  # <- clave para evitar corrupción OOXML
                color="90CAF9"
            )
        )
    except Exception:
        # Si la versión de Excel/openpyxl no soporta algo, lo omitimos sin romper el archivo.
        pass

    # Ajuste de columnas y congelar encabezados
    widths = {"A": 26, "B": 16, "C": 28, "D": 2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"  # deja titulo visible

    # --- Hoja de Evolución ---
    ws2 = wb.create_sheet("Evolución")
    ws2.append(["Fecha", "Acciones"])
    ws2["A1"].font = ws2["B1"].font = bold
    ws2["A1"].fill = ws2["B1"].fill = head_fill
    ws2["A1"].border = ws2["B1"].border = box

    for p in ts:
        ws2.append([p.get("date"), p.get("count", 0)])

    # Formato y anchos
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    for r in range(2, ws2.max_row + 1):
        ws2[f"B{r}"].number_format = "0"
        ws2[f"B{r}"].alignment = right
        ws2[f"A{r}"].alignment = left
        ws2[f"A{r}"].border = ws2[f"B{r}"].border = box
    ws2.freeze_panes = "A2"

    # Gráfica solo si hay datos (≥ 1 fila)
    if ws2.max_row >= 3:
        try:
            chart = LineChart()
            chart.title = "Acciones por día"
            chart.style = 2
            chart.y_axis.title = "Acciones"
            chart.y_axis.number_format = "0"
            chart.x_axis.title = "Fecha"

            data = Reference(ws2, min_col=2, min_row=1, max_col=2, max_row=ws2.max_row)
            cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width  = 22
            ws2.add_chart(chart, "D2")
        except Exception:
            # Si el Excel del usuario no soporta algo del chart, no rompemos el archivo.
            pass

    # --- salida ---
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f'attachment; filename="summary_{days}d.xlsx"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


@bp.get("/api/reports/summary.pdf")
def report_summary_pdf():
    """
    PDF ejecutivo (KPIs + evolución) con diseño simple y profesional.
    - Encabezado con rango de fechas y fecha de generación
    - 4 tarjetas KPI (Total, OK, Fallidas, Tasa éxito)
    - Gráfica de barras con la evolución diaria
    - Tabla con las últimas 30 filas (o las que entren) y paginación
    """
    if canvas is None:
        return jsonify(ok=False, error="reportlab no instalado"), 500

    # ------------------ Parámetros ------------------
    try:
        days = int(request.args.get("days", "30"))
    except ValueError:
        days = 30

    end_dt = datetime.utcnow()
    start_dt = end_dt - timedelta(days=days)

    # ------------------ Datos ------------------
    with SessionLocal() as db:
        kpis = build_kpis(db, start_dt, end_dt) or {}
        ts = build_timeseries(db, start_dt, end_dt) or []

    # Evita crashear si ts viene vacío
    if not isinstance(ts, list):
        ts = []

    # ------------------ Helpers de dibujo ------------------
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm

    PAGE_W, PAGE_H = A4
    MARGIN = 18 * mm

    def _header(c, page=1, pages=1):
        c.setFont("Helvetica-Bold", 15)
        c.setFillColor(colors.black)
        c.drawString(MARGIN, PAGE_H - MARGIN + 4*mm, "Resumen KPIs — Operación")
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.grey)
        c.drawString(MARGIN, PAGE_H - MARGIN - 1*mm,
                     f"Ventana: {start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}  |  Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")

        # Línea sutil
        c.setStrokeColor(colors.lightgrey)
        c.setLineWidth(0.5)
        c.line(MARGIN, PAGE_H - MARGIN - 5*mm, PAGE_W - MARGIN, PAGE_H - MARGIN - 5*mm)

    def _footer(c, page=1):
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(PAGE_W - MARGIN, MARGIN - 4*mm, f"Página {page}")

    def kpi_card(c, x, y, w, h, title, value, subtitle="", accent=colors.HexColor("#2563eb")):
        # Fondo
        c.setFillColor(colors.HexColor("#f8fafc"))
        c.setStrokeColor(colors.HexColor("#e5e7eb"))
        c.setLineWidth(0.6)
        c.roundRect(x, y, w, h, 6, stroke=1, fill=1)

        # Barra acento
        c.setFillColor(accent)
        c.roundRect(x, y + h - 4, w, 4, 2, stroke=0, fill=1)

        # Texto
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawString(x + 6, y + h - 14, title)

        c.setFont("Helvetica-Bold", 16)
        c.drawString(x + 6, y + h - 36, f"{value}")

        if subtitle:
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.grey)
            c.drawString(x + 6, y + 8, subtitle)

    def bar_chart(c, x, y, w, h, series, title="Evolución diaria"):
        """
        series: [{'date': 'YYYY-MM-DD', 'count': int}, ...]
        Dibuja un pequeño bar chart auto-escalado.
        """
        # Marco
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor("#e5e7eb"))
        c.setLineWidth(0.6)
        c.roundRect(x, y, w, h, 6, stroke=1, fill=1)

        # Título
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.black)
        c.drawString(x + 8, y + h - 16, title)

        inner_x = x + 10
        inner_y = y + 16
        inner_w = w - 20
        inner_h = h - 32

        # Datos
        counts = [int(s.get("count", 0) or 0) for s in series]
        maxv = max(counts) if counts else 1
        n = len(series)
        if n == 0:  # “No data”
            c.setFont("Helvetica", 9)
            c.setFillColor(colors.grey)
            c.drawCentredString(x + w/2, y + h/2 - 4, "Sin datos")
            return

        gap = 1  # separación mínima
        bar_w = max(1.0, (inner_w / max(n, 1)) - gap)
        c.setFillColor(colors.HexColor("#6366f1"))  # Indigo

        for i, s in enumerate(series):
            val = int(s.get("count", 0) or 0)
            bar_h = 0 if maxv == 0 else (val / maxv) * inner_h
            bx = inner_x + i * (bar_w + gap)
            by = inner_y
            # Barra
            c.rect(bx, by, bar_w, bar_h, stroke=0, fill=1)

        # Ejes sutiles
        c.setStrokeColor(colors.HexColor("#e5e7eb"))
        c.setLineWidth(0.5)
        c.line(inner_x, inner_y, inner_x + inner_w, inner_y)  # eje X

        # Etiqueta max
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.grey)
        c.drawRightString(inner_x + inner_w, inner_y + inner_h + 2, f"max: {maxv}")

    def table_page(c, x, y, w, h, rows, start_idx=0, rows_per_page=28, title="Detalle últimos días"):
        """
        Dibuja una tabla simple con fecha y acciones.
        Devuelve el siguiente índice (para paginación).
        """
        c.setFillColor(colors.white)
        c.setStrokeColor(colors.HexColor("#e5e7eb"))
        c.setLineWidth(0.6)
        c.roundRect(x, y, w, h, 6, stroke=1, fill=1)

        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.black)
        c.drawString(x + 8, y + h - 16, title)

        # Cabecera
        col1_w = 90 * mm
        col2_w = w - 20 - col1_w
        tx = x + 10
        ty = y + h - 32

        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(colors.HexColor("#6b7280"))
        c.drawString(tx, ty, "Fecha")
        c.drawRightString(x + w - 10, ty, "Acciones")

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.black)

        i = 0
        idx = start_idx
        line_h = 11
        ty -= 8
        while idx < len(rows) and i < rows_per_page and (ty - line_h) > (y + 10):
            r = rows[idx]
            date = str(r.get("date"))
            cnt = int(r.get("count", 0) or 0)
            ty -= line_h
            c.drawString(tx, ty, date)
            c.drawRightString(x + w - 10, ty, f"{cnt}")
            i += 1
            idx += 1

        return idx

    # ------------------ Render ------------------
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)

    page = 1
    _header(c, page)
    _footer(c, page)

    # Layout de la primera página
    # Tarjetas KPI (grid 2x2)
    card_w = (PAGE_W - 2*MARGIN - 12*mm) / 2
    card_h = 24 * mm
    card_x1 = MARGIN
    card_x2 = MARGIN + card_w + 12*mm
    y_top = PAGE_H - MARGIN - 18*mm

    # Valores KPI
    total = int(kpis.get("total_actions", 0) or 0)
    ok = int(kpis.get("ok", 0) or 0)
    failed = int(kpis.get("failed", 0) or 0)
    rate = float(kpis.get("success_rate", 0) or 0.0)

    kpi_card(c, card_x1, y_top - card_h, card_w, card_h, "Total acciones",
             f"{total:,}".replace(",", ","), "Suma de acciones", colors.HexColor("#0ea5e9"))
    kpi_card(c, card_x2, y_top - card_h, card_w, card_h, "OK",
             f"{ok:,}".replace(",", ","), "Completadas", colors.HexColor("#22c55e"))

    kpi_card(c, card_x1, y_top - 2*card_h - 8*mm, card_w, card_h, "Fallidas",
             f"{failed:,}".replace(",", ","), "Errores/timeout", colors.HexColor("#ef4444"))
    kpi_card(c, card_x2, y_top - 2*card_h - 8*mm, card_w, card_h, "Tasa de éxito",
             f"{rate:.2f}%", "OK / Total", colors.HexColor("#6366f1"))

    # Gráfica barras
    chart_y = y_top - 2*card_h - 8*mm - 10*mm - 60*mm
    if chart_y < MARGIN + 60*mm:
        chart_y = MARGIN + 60*mm
    bar_chart(c, MARGIN, chart_y, PAGE_W - 2*MARGIN, 60*mm, ts, "Evolución de acciones")

    # Tabla detalle (paginada)
    idx = table_page(c, MARGIN, MARGIN + 8*mm, PAGE_W - 2*MARGIN, chart_y - (MARGIN + 16*mm), ts,
                     start_idx=0, rows_per_page=28, title="Detalle por día")

    # Páginas adicionales si faltan filas
    while idx < len(ts):
        c.showPage()
        page += 1
        _header(c, page)
        _footer(c, page)
        idx = table_page(c, MARGIN, MARGIN + 8*mm, PAGE_W - 2*MARGIN, PAGE_H - 2*MARGIN - 20*mm,
                         ts, start_idx=idx, rows_per_page=40, title="Detalle por día (cont.)")

    c.showPage()
    c.save()

    # ------------------ Respuesta ------------------
    bio.seek(0)
    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = f'attachment; filename="summary_{days}d.pdf"'
    resp.headers["Cache-Control"] = "no-store"
    return resp



# =============================================================================
# Tableros PDF/HTML (opcionales para dirección/forecast)
# =============================================================================

def _css_base():
    """CSS minimalista para los HTML/PDF de board/forecast."""
    return """
    <style>
      body{font-family:Inter,Arial,sans-serif;color:#111;background:#fff;margin:0;padding:24px}
      h1{font-size:20px;font-weight:800;margin:0 0 8px}
      .muted{color:#6b7280;font-size:12px;margin:0 0 12px}
      .card{border:1px solid #e5e7eb;border-radius:12px;padding:16px;margin-top:12px}
      .kpi{display:grid;grid-template-columns:1fr 1fr;gap:12px}
      .kpi .box{background:#f8fafc;border:1px solid #e5e7eb;border-radius:10px;padding:12px}
      .kpi .label{color:#6b7280;font-size:12px}
      .kpi .value{font-size:18px;font-weight:700;margin-top:4px}
      table{width:100%;border-collapse:collapse;margin-top:10px}
      th,td{border:1px solid #e5e7eb;padding:8px;text-align:left;font-size:12px}
      th{background:#f3f4f6}
    </style>
    """

def _render_pdf_or_html(html: str, filename: str = "reporte.pdf"):
    """Intenta renderizar HTML→PDF con weasyprint; si falla, regresa HTML."""
    try:
        from weasyprint import HTML  # optional
        pdf_bytes = HTML(string=html).write_pdf()
        return Response(
            pdf_bytes,
            headers={
                "Content-Type": "application/pdf",
                "Content-Disposition": f'inline; filename="{filename}"',
            },
        )
    except Exception:
        return Response(html, headers={"Content-Type": "text/html; charset=utf-8"})


@bp.get("/api/reports/board.pdf")
def api_report_board_pdf():
    """
    PDF Ejecutivo para Dirección (siempre PDF con ReportLab).
    Diseño: tarjetas KPI + barra de aging + Top 5 clientes.
    """
    if canvas is None:
        return jsonify(ok=False, error="reportlab no instalado"), 500

    # Parámetro opcional ?days= para referencia visual; los datos
    # se traen 'al día' del ERP (no se filtra por days aquí).
    try:
        days = int(request.args.get("days", "30"))
    except Exception:
        days = 30

    rows = _load_invoices_for_board(limit=2000)
    M = _board_metrics(rows)

    # ======= Armado del PDF =======
    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    W, H = A4

    def money(x): 
        return f"${x:,.2f}"

    # Márgenes
    left, right, top, bottom = 40, W - 40, H - 40, 40
    y = top

    # Title
    c.setFillColor(PALETTE["ink"])
    c.setFont("Helvetica-Bold", 16)
    c.drawString(left, y, "Vista Ejecutiva — Board Directivo")
    c.setFont("Helvetica", 9)
    c.setFillColor(PALETTE["muted"])
    c.drawString(left, y - 14, f"Horizonte en UI: últimos {days} días • Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    y -= 36

    # Helper: tarjeta KPI
    def kpi_card(x, y, w, h, label, value, sub=None, stripe=PALETTE["blue"]):
        c.setFillColor(PALETTE["card"]); c.setStrokeColor(PALETTE["line"])
        c.roundRect(x, y - h, w, h, 8, stroke=1, fill=1)
        # stripe superior
        c.setFillColor(stripe); c.roundRect(x, y, w, 4, 2, stroke=0, fill=1)

        c.setFillColor(PALETTE["muted"]); c.setFont("Helvetica", 9)
        c.drawString(x + 10, y - 18, label)

        c.setFillColor(PALETTE["ink"]); c.setFont("Helvetica-Bold", 16)
        c.drawString(x + 10, y - 40, value)

        if sub:
            c.setFillColor(PALETTE["muted"]); c.setFont("Helvetica", 8)
            c.drawString(x + 10, y - 58, sub)

    # Grid de 4 KPI
    gap = 14
    card_w = (right - left - (gap * 3)) / 4
    card_h = 70
    Xs = [left + i * (card_w + gap) for i in range(4)]
    y_top = y

    kpi_card(Xs[0], y_top, card_w, card_h,
             "Total Cartera", money(M["total"]), "monto total",
             stripe=PALETTE["slate"])
    kpi_card(Xs[1], y_top, card_w, card_h,
             "En Tiempo (verde + amarillo)", money(M["en_tiempo"]),
             "a tiempo + vence hoy", stripe=PALETTE["green"])
    kpi_card(Xs[2], y_top, card_w, card_h,
             "Atraso leve (1–14d)", money(M["late_1_14"]), "naranja",
             stripe=PALETTE["orange"])
    kpi_card(Xs[3], y_top, card_w, card_h,
             "Riesgo alto (≥15d)", money(M["riesgo_alto"]), "rojo",
             stripe=PALETTE["red"])

    y = y_top - (card_h + 26)

    # ===== Barra de distribución (aging) =====
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(PALETTE["ink"])
    c.drawString(left, y, "Distribución por riesgo y vencimiento")
    y -= 10
    c.setStrokeColor(PALETTE["line"]); c.line(left, y, right, y)
    y -= 16

    total = M["total"] or 1.0
    seg_w = (right - left)
    bar_h = 16

    segments = [
        ("A tiempo", max(M["on_time"], 0.0), PALETTE["green"]),
        ("Vence hoy", max(M["due_today"], 0.0), PALETTE["yellow"]),
        ("1–14d", max(M["late_1_14"], 0.0), PALETTE["orange"]),
        ("15–30d", max(M["late_15_30"], 0.0), PALETTE["orange"]),
        ("31–60d", max(M["late_31_60"], 0.0), PALETTE["orange"]),
        ("60+d", max(M["late_60_plus"], 0.0), PALETTE["red"]),
    ]

    x = left
    for label, val, col in segments:
        w = seg_w * (val / total)
        if w < 0.5:
            continue
        c.setFillColor(col)
        c.rect(x, y - bar_h, w, bar_h, stroke=0, fill=1)
        x += w

    # Leyenda
    y -= (bar_h + 10)
    c.setFont("Helvetica", 9)
    x = left
    for label, val, col in segments:
        c.setFillColor(col); c.rect(x, y - 8, 10, 10, stroke=0, fill=1)
        c.setFillColor(PALETTE["ink"]); c.drawString(x + 14, y - 0, f"{label}: {money(val)}")
        x += 150

    y -= 24

    # ===== Top 5 clientes =====
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(PALETTE["ink"])
    c.drawString(left, y, "Top 5 Clientes por saldo")
    y -= 12

    c.setFillColor(PALETTE["card"]); c.setStrokeColor(PALETTE["line"])
    c.roundRect(left, y - 100, right - left, 100, 8, stroke=1, fill=1)

    c.setFont("Helvetica", 10)
    c.setFillColor(PALETTE["ink"])
    yy = y - 16
    for i, row in enumerate(M["top5"], start=1):
        c.drawString(left + 12, yy, f"{i}. {row['cliente']}")
        c.drawRightString(right - 12, yy, money(row["monto"]))
        yy -= 18

    # Pie
    c.setFont("Helvetica", 8)
    c.setFillColor(PALETTE["muted"])
    c.drawRightString(right, bottom, "Generado automáticamente — Motor de Cobranza")

    c.showPage()
    c.save()

    bio.seek(0)
    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/pdf"
    resp.headers["Content-Disposition"] = 'attachment; filename="board_directivo.pdf"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


@bp.get("/api/reports/board.xlsx")
def api_report_board_xlsx():
    """
    Excel Ejecutivo — Board Directivo
    - Hoja 'Board': KPIs + gráfico de aging (0-14, 15-30, 31-60, 60+)
    - Hoja 'Detalle': listado de facturas (cliente, vencimiento, DPD, saldo)
    """
    from io import BytesIO
    from datetime import datetime, date
    try:
        # Carga directa desde ERP
        from erp_client import list_unpaid_invoices
    except Exception:
        return jsonify(ok=False, error="No se pudo importar erp_client.list_unpaid_invoices"), 500

    # --- Dependencias openpyxl ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
    except Exception:
        return jsonify(ok=False, error="openpyxl no instalado"), 500

    # ---------- Helpers locales ----------
    def _safe_float(x):
        try:
            return float(x or 0)
        except Exception:
            return 0.0

    def _days_past_due(iso_date: str) -> int:
        if not iso_date:
            return 0
        try:
            d = date.fromisoformat(iso_date[:10])
        except Exception:
            return 0
        return max(0, (date.today() - d).days)

    def _autosize(ws, cols):
        for c in cols:
            ws.column_dimensions[get_column_letter(c)].auto_size = True
            # Un ancho razonable si el auto_size del viewer no aplica:
            ws.column_dimensions[get_column_letter(c)].width = ws.column_dimensions[get_column_letter(c)].width or 18

    # ---------- Datos desde ERP ----------
    try:
        rows = list_unpaid_invoices(limit=5000, include_draft=False, order_by="due_date asc")
    except Exception as e:
        return jsonify(ok=False, error=f"ERP error: {e}"), 502

    # ---------- Métricas ----------
    total = 0.0
    due_today = 0.0
    en_tiempo = 0.0          # a tiempo + vence hoy (DPD == 0)
    late_1_14 = 0.0
    late_15_30 = 0.0
    late_31_60 = 0.0
    late_60_plus = 0.0

    # Top clientes por saldo
    by_client = {}

    for r in rows:
        amt = _safe_float(r.get("outstanding_amount"))
        total += amt
        cli = (r.get("customer") or "").strip() or "Cliente"
        by_client[cli] = by_client.get(cli, 0.0) + amt

        dpd = _days_past_due(r.get("due_date"))
        if dpd == 0:
            en_tiempo += amt
            # si justo vence hoy (misma condición en este dataset)
            due_today += amt
        elif 1 <= dpd <= 14:
            late_1_14 += amt
        elif 15 <= dpd <= 30:
            late_15_30 += amt
        elif 31 <= dpd <= 60:
            late_31_60 += amt
        else:  # 61+
            late_60_plus += amt

    riesgo_alto = late_15_30 + late_31_60 + late_60_plus
    top5 = sorted(by_client.items(), key=lambda kv: kv[1], reverse=True)[:5]

    # ---------- Libro / estilos ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Board"

    money_fmt = '#,##0.00'
    f_title = Font(bold=True, size=14)
    f_h2    = Font(bold=True, size=12)
    f_bold  = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right  = Alignment(horizontal="right",  vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    box  = Border(top=thin, left=thin, right=thin, bottom=thin)
    fill_header = PatternFill("solid", fgColor="EEF2FF")  # lila claro
    fill_card   = PatternFill("solid", fgColor="F8FAFC")  # gris muy claro

    # ---------- Título ----------
    ws.merge_cells("A1:F1")
    ws["A1"] = "Vista Ejecutiva — Board Directivo"
    ws["A1"].font = f_title
    ws["A1"].alignment = align_center
    ws["A2"] = "Generado"
    ws["B2"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    # ---------- KPIs (tabla) ----------
    ws.append([])
    ws.append(["KPI", "Valor", "Descripción"])
    hdr = ws.max_row
    ws[f"A{hdr}"].font = f_bold
    ws[f"B{hdr}"].font = f_bold
    ws[f"C{hdr}"].font = f_bold
    for c in ("A", "B", "C"):
        ws[f"{c}{hdr}"].fill = fill_header
        ws[f"{c}{hdr}"].border = box
        ws[f"{c}{hdr}"].alignment = Alignment(vertical="center")

    data_kpi = [
        ("Total Cartera",              total,         "Monto total"),
        ("En Tiempo (verde+amarillo)", en_tiempo,     "A tiempo + vence hoy"),
        ("Atraso leve (1–14d)",        late_1_14,     "Naranja"),
        ("Riesgo alto (≥15d)",         riesgo_alto,   "Rojo"),
        ("Vence hoy",                  due_today,     ""),
        ("60+ días",                   late_60_plus,  ""),
    ]
    for k, v, d in data_kpi:
        ws.append([k, v, d])
        r = ws.max_row
        ws[f"A{r}"].font = f_bold
        ws[f"B{r}"].number_format = money_fmt
        for c in ("A", "B", "C"):
            ws[f"{c}{r}"].border = box
        ws[f"A{r}"].fill = fill_card

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 36

    # ---------- Aging (tabla + gráfico) ----------
    ws.append([])
    ws.append(["Aging (riesgo por vencimiento)"])
    ws[f"A{ws.max_row}"].font = f_h2

    ws.append(["Tramo", "Monto"])
    hdr2 = ws.max_row
    for c in ("A", "B"):
        ws[f"{c}{hdr2}"].font = f_bold
        ws[f"{c}{hdr2}"].fill = fill_header
        ws[f"{c}{hdr2}"].border = box

    buckets = {
        "0-14": late_1_14,
        "15-30": late_15_30,
        "31-60": late_31_60,
        "60+": late_60_plus
    }

    first_row = hdr2 + 1
    for tramo in ("0-14", "15-30", "31-60", "60+"):
        ws.append([tramo, float(buckets.get(tramo, 0.0))])
        r = ws.max_row
        ws[f"B{r}"].number_format = money_fmt
        ws[f"A{r}"].fill = fill_card
        ws[f"A{r}"].border = box
        ws[f"B{r}"].border = box

    _autosize(ws, [1, 2])

    last_row = ws.max_row
    vals_ref = Reference(ws, min_col=2, min_row=first_row, max_col=2, max_row=last_row)
    cats_ref = Reference(ws, min_col=1, min_row=first_row, max_row=last_row)
    max_val = max((_safe_float(ws.cell(row=r, column=2).value) for r in range(first_row, last_row + 1)), default=0.0)

    if max_val > 0:
        chart = BarChart()
        chart.title = "Distribución por tramo"
        chart.add_data(vals_ref, titles_from_data=False)  # una sola serie (montos)
        chart.set_categories(cats_ref)                    # categorías = tramos
        chart.varyColors = True
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.y_axis.title = "Monto"
        chart.x_axis.title = "Tramo (días)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = float(max_val) * 1.2
        chart.height = 9
        chart.width  = 20
        # Ancla el gráfico a la derecha de la tabla de KPIs (cámbialo si lo quieres en otra posición)
        ws.add_chart(chart, "E5")
    else:
        # Mensaje en lugar del gráfico
        anchor_row = hdr2 - 1 if hdr2 > 2 else hdr2
        ws.merge_cells(f"E{anchor_row}:I{anchor_row+6}")
        cell = ws[f"E{anchor_row}"]
        cell.value = "Sin datos para graficar (todos los tramos están en $0)."
        cell.alignment = align_center
        cell.fill = fill_card
        cell.border = box

    # ---------- Detalle ----------
    ws2 = wb.create_sheet("Detalle")
    ws2.append(["Cliente", "Fecha Venc.", "DPD", "Saldo"])
    for c in ("A", "B", "C", "D"):
        ws2[f"{c}1"].font = f_bold
        ws2[f"{c}1"].fill = fill_header
        ws2[f"{c}1"].border = box

    for r in rows:
        dpd = _days_past_due(r.get("due_date"))
        amt = _safe_float(r.get("outstanding_amount"))
        ws2.append([r.get("customer"), r.get("due_date"), dpd, amt])

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 16
    for r in range(2, ws2.max_row + 1):
        ws2[f"C{r}"].alignment = align_right
        ws2[f"D{r}"].number_format = money_fmt

    ws.freeze_panes = "A5"   # congela encabezado en Board
    ws2.freeze_panes = "A2"  # congela encabezado en Detalle

    # ---------- Respuesta ----------
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = 'attachment; filename="board_directivo.xlsx"'
    resp.headers["Cache-Control"] = "no-store"
    return resp


@bp.get("/api/reports/forecast.pdf")
def api_report_forecast_pdf():
    """Predicción de cobros por fecha de vencimiento (simple) en horizonte."""
    if not callable(list_unpaid_invoices):
        return jsonify(ok=False, error="erp_client.list_unpaid_invoices no disponible"), 500

    horizon = int(request.args.get("horizon", "30"))
    rows = list_unpaid_invoices(limit=500, include_draft=False)
    today = dt.date.today()

    buckets = {}  # date -> amount
    for r in rows:
        dd = r.get("due_date")
        try:
            d = dt.date.fromisoformat(dd) if dd else None
        except Exception:
            d = None
        amt = float(r.get("outstanding_amount") or 0)
        if d and 0 <= (d - today).days <= horizon:
            buckets[d.isoformat()] = buckets.get(d.isoformat(), 0.0) + amt

    total_expected = sum(buckets.values())
    rows_html = "".join(f"<tr><td>{k}</td><td>${v:,.2f}</td></tr>" for k, v in sorted(buckets.items()))
    html = f"""
    <!doctype html><html><head><meta charset="utf-8">{_css_base()}</head><body>
      <h1>Predicción de Cobros</h1>
      <p class="muted">Horizonte: {horizon} días</p>
      <div class="card"><b>Total esperado:</b> ${total_expected:,.2f}</div>
      <div class="card">
        <table>
          <thead><tr><th>Día</th><th>Monto esperado</th></tr></thead>
          <tbody>{rows_html or '<tr><td colspan="2">Sin predicción disponible</td></tr>'}</tbody>
        </table>
      </div>
    </body></html>
    """
    return _render_pdf_or_html(html, filename="forecast.pdf")



router = APIRouter()

ALLOWED_FORMATS = {"pdf", "xlsx"}

def split_csv(value: Optional[str | List[str]]) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [p.strip() for p in str(value).split(",") if p.strip()]

class ReportParams(BaseModel):
    days: int = Field(30, ge=1, le=365)
    include_draft: bool = False
    formats: List[str] = Field(default_factory=lambda: ["pdf", "xlsx"])
    groups: List[str] = Field(default_factory=list)
    categories: List[str] = Field(default_factory=lambda: ["verde","amarillo","naranja","rojo"])
    emails: List[EmailStr] = Field(default_factory=list)
    min_dpd: int = Field(0, ge=0)

    @validator("formats", pre=True)
    def parse_formats(cls, v):
        vals = split_csv(v)
        vals = [s.lower() for s in vals]
        # valida formatos permitidos
        bad = [x for x in vals if x not in ALLOWED_FORMATS]
        if bad:
            raise ValueError(f"formatos no permitidos: {bad}")
        # de-dup
        return list(dict.fromkeys(vals))

    @validator("groups", pre=True)
    def parse_groups(cls, v): return split_csv(v)

    @validator("categories", pre=True)
    def parse_categories(cls, v): return split_csv(v)

    @validator("emails", pre=True)
    def parse_emails(cls, v): return split_csv(v)

@router.post("/api/report-schedule/run")
def run_report_post(payload: ReportParams = Body(...)):
    # payload llega tipado y validado
    return {"ok": True, "params": payload.dict()}

@router.get("/api/report-schedule/run")
def run_report_get(
    days: int = Query(30, ge=1, le=365),
    include_draft: bool = Query(False),
    formats: Optional[str] = Query(None, description="csv o repetir ?formats=pdf&formats=xlsx"),
    groups: Optional[str] = Query(None),
    categories: Optional[str] = Query("verde,amarillo,naranja,rojo"),
    emails: Optional[str] = Query(None),
    min_dpd: int = Query(0, ge=0),
):
    # Reusa el modelo para normalizar GET también
    params = ReportParams(
        days=days,
        include_draft=include_draft,
        formats=formats or ["pdf","xlsx"],
        groups=groups,
        categories=categories,
        emails=emails,
        min_dpd=min_dpd,
    )
    return {"ok": True, "params": params.dict()}


@bp.post("/api/report-schedule/test-send")
def api_report_test_send():
    body = request.get_json(silent=True) or {}

    period_days   = int(body.get("period_days", 30))
    include_draft = bool(body.get("include_draft", False))
    subject = body.get("subject") or f"Reporte CxC — {dt.date.today().isoformat()}"

    recipients = set()

    # 1) emails directos (ya resueltos)
    for e in (body.get("emails") or []):
        if isinstance(e, str) and "@" in e:
            recipients.add(e.strip())

    # 2) extra_emails
    for e in (body.get("extra_emails") or []):
        if isinstance(e, str) and "@" in e:
            recipients.add(e.strip())

    # 3) groups_obj: { direccion: ["a@..."], cobranza: [...] }
    groups_obj = body.get("groups_obj")
    if not groups_obj and isinstance(body.get("groups"), dict):
        groups_obj = body["groups"]
    if isinstance(groups_obj, dict):
        for arr in groups_obj.values():
            for e in (arr or []):
                if isinstance(e, str) and "@" in e:
                    recipients.add(e.strip())

    # 4) groups como array de slugs -> resolver en servidor
    if not recipients:
        groups = body.get("groups") or []
        if isinstance(groups, list) and groups:
            store = _load_groups()  # tu función que lee los grupos desde config/DB
            for g in groups:
                key = _slugify(g)
                for e in store.get(key, []):
                    if isinstance(e, str) and "@" in e:
                        recipients.add(e.strip())

    recipients = sorted(recipients)

    if not recipients:
        return jsonify({
            "ok": False,
            "error": "NO_RECIPIENTS",
            "hint": "Envía emails[], extra_emails[], groups[] (slugs) o groups_obj{} con correos."
        }), 400

    links = _report_links(period_days)

    html = body.get("html")
    if not html:
        html = f"""<!doctype html>
<html lang="es"><head><meta charset="utf-8">
<style>
  body{{font-family:Inter,Arial,sans-serif;color:#111}}
  .btn{{display:inline-block;padding:10px 14px;border-radius:10px;
       background:#2563eb;color:#fff;text-decoration:none;font-weight:600;margin-right:8px}}
  .small{{color:#6b7280;font-size:12px}}
</style></head>
<body>
  <h1>Reporte de Cartera (CxC)</h1>
  <p>Rango analizado: últimos <b>{period_days}</b> días</p>
  <p>
    <a class="btn" href="{links['pdf']}">Descargar PDF</a>
    <a class="btn" href="{links['xlsx']}">Descargar XLSX</a>
  </p>
  <p class="small">Enviado automáticamente por el sistema.</p>
</body></html>"""

    payload = {
        "emails": recipients,
        "subject": subject,
        "html": html,
        "links": links
    }

    if not EMAIL_WEBHOOK_URL:
        # Modo debug
        return jsonify({"ok": True, "debug": "NO_WEBHOOK_CONFIGURED", "payload": payload})

    try:
        r = requests.post(EMAIL_WEBHOOK_URL, json=payload, timeout=20)
        ok = 200 <= r.status_code < 300
        ctype = r.headers.get("Content-Type", "")
        resp = r.json() if "application/json" in ctype else r.text
        return jsonify({"ok": ok, "status": r.status_code, "response": resp})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
